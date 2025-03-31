import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

from fastapi import FastAPI, Form, File, UploadFile  # type: ignore
from fastapi.responses import HTMLResponse  # type: ignore
from fastapi.middleware.cors import CORSMiddleware  # type: ignore
import os
import openpyxl  # type: ignore
from processing import fetch_answer
import re
import stat
import json
import base64
from io import BytesIO
from PIL import Image
import httpx  # type: ignore
import aiofiles
from typing import List
from git_api import GA1_13, GA2_3, GA2_7, GA4_8, GA2_9_file, GA2_6_file

app = FastAPI()  # Ensure this is defined early

# CORS Configuration (Vercel allows any origin by default)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "tasks.xlsx")

def load_tasks_from_excel():
    if not os.path.exists(EXCEL_FILE):
        logger.error(f"File {EXCEL_FILE} does not exist.")
        return {}, {}
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        tasks = {row[0]: row[1] for row in sheet.iter_rows(
            min_row=2, values_only=True) if row[0] and row[1]}

        tasks_answers = {row[0]: row[2] for row in sheet.iter_rows(
            min_row=2, values_only=True) if row[0] and row[2]}
        workbook.close()
        return tasks, tasks_answers
    except Exception as e:
        logger.error(f"Error loading tasks from {EXCEL_FILE}: {e}")
        return {}, {}

TASKS, TASKS_ANSWERS = load_tasks_from_excel()

# ... (rest of your code)
